"""export 生成 Excel 测试。"""

from fapiao.export import export_excel
from fapiao.models import InvoiceRecord, PendingItem
from fapiao.store import Store


def test_export_creates_two_sheets(tmp_path):
    s = Store(tmp_path / "db.sqlite", tmp_path / "archive")
    s.save_invoice(InvoiceRecord(发票号码="INV1", 开票日期="2026-06-01", 销售方名称="餐厅",
                                 价税合计=100.0, 消费明细=[{"名称": "餐饮服务"}], confidence=0.9))
    s.add_pending(PendingItem(email_uid=2, subject="发票", sender="x@y.com", reason="需登录"))

    out = export_excel(s, tmp_path / "out.xlsx")
    assert out.exists()
    s.close()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == ["发票汇总", "待处理"]

    ws = wb["发票汇总"]
    assert ws["A1"].value == "开票日期"
    # 第二行是数据
    headers = [c.value for c in ws[1]]
    row2 = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(headers)}
    assert row2["发票号码"] == "INV1"
    assert row2["消费明细"] == "餐饮服务"   # 明细被压成可读文本

    ws2 = wb["待处理"]
    assert ws2.cell(row=2, column=4).value == "需登录"


def test_export_empty_db(tmp_path):
    s = Store(tmp_path / "db.sqlite", tmp_path / "archive")
    out = export_excel(s, tmp_path / "out.xlsx")
    assert out.exists()
    s.close()

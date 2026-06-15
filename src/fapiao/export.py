"""导出:从 SQLite 生成 Excel 汇总(发票汇总 + 待处理 两张表)。"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .store import Store

# (表头, 取数的列名)
INVOICE_COLUMNS = [
    ("开票日期", "开票日期"),
    ("发票号码", "发票号码"),
    ("发票代码", "发票代码"),
    ("发票类型", "发票类型"),
    ("销售方名称", "销售方名称"),
    ("销售方税号", "销售方税号"),
    ("购买方名称", "购买方名称"),
    ("购买方税号", "购买方税号"),
    ("金额(不含税)", "金额"),
    ("税率", "税率"),
    ("税额", "税额"),
    ("价税合计", "价税合计"),
    ("消费明细", "消费明细"),
    ("备注", "备注"),
    ("状态", "status"),
    ("置信度", "confidence"),
    ("归档文件", "archive_path"),
]

PENDING_COLUMNS = [
    ("时间", "created_at"),
    ("邮件主题", "subject"),
    ("发件人", "sender"),
    ("原因", "reason"),
    ("链接", "link"),
]

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _summarize_明细(raw: str) -> str:
    """把消费明细 JSON 压成一行可读文本。"""
    try:
        items = json.loads(raw) if raw else []
    except (json.JSONDecodeError, TypeError):
        return raw or ""
    names = [str(it.get("名称", "")).strip() for it in items if isinstance(it, dict)]
    names = [n for n in names if n]
    return "; ".join(names)


def _write_sheet(ws, columns, rows, transform=None):
    headers = [c[0] for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        values = []
        for _, key in columns:
            val = row[key] if key in row.keys() else ""
            if transform:
                val = transform(key, val)
            values.append(val)
        ws.append(values)

    # 简单列宽
    for i, _ in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "A2"


def export_excel(store: Store, output_file: str | Path) -> Path:
    """生成 Excel 文件,返回路径。"""
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"

    def inv_transform(key, val):
        if key == "消费明细":
            return _summarize_明细(val)
        return val if val is not None else ""

    _write_sheet(ws, INVOICE_COLUMNS, store.all_invoices(), inv_transform)

    ws2 = wb.create_sheet("待处理")
    _write_sheet(ws2, PENDING_COLUMNS, store.all_pending(),
                 lambda k, v: v if v is not None else "")

    wb.save(output_file)
    return output_file
